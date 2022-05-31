from converter import app
import os
import io
from flask import render_template, request, flash, redirect, send_file
from werkzeug.utils import secure_filename
from docxtpl import DocxTemplate
from openpyxl import load_workbook

ALLOWED_EXTENSIONS_FOR_DOCX = {"docx"}
ALLOWED_EXTENSIONS_FOR_XLSX = {"xlsx"}


def convert(excel, word):
    wb = load_workbook(excel)
    sheet = wb.active
    last_row = sheet.max_row
    dict1 = {}
    for row in wb.worksheets[0].iter_rows(min_row=2, max_row=last_row, max_col=2):
        dict1[row[0].value] = f"({row[1].value})"

    doc = DocxTemplate(word)
    context = dict1

    doc.render(context)
    doc.save(os.path.join(app.config["UPLOAD_FOLDER"], "Template_Rendered.docx"))


def allowed_file_docx(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS_FOR_DOCX
    )


def allowed_file_xlsx(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS_FOR_XLSX
    )


@app.route("/")
def homeView():
    return render_template("main_page.html")


@app.route("/", methods=["POST"])
def upload_files():
    if "word" not in request.files:
        flash("No file part")
        return redirect(request.url)
    word_file = request.files["word"]
    excel_file = request.files["excel"]
    if word_file.filename == "" or excel_file.filename == "":
        flash("No files selected for uploading")
        return redirect(request.url)
    if excel_file and allowed_file_xlsx(excel_file.filename) and word_file and allowed_file_docx(word_file.filename):
        secure_filename(excel_file.filename)
        secure_filename(word_file.filename)

    else:
        flash("Wrong types of files")
        return redirect(request.url)

    convert(excel_file, word_file)
    return redirect("download_and_delete")


@app.route("/download_and_delete")
def download_and_delete():
    # fetch file contents
    current_dirname = os.path.dirname(os.path.abspath(__file__))
    filepath = f"{current_dirname}/static/files/Template_Rendered.docx"
    file_content = open(filepath, "rb").read()
    # write the file into in-memory file-like io object
    return_file = io.BytesIO(file_content)
    # delete the file
    os.remove(filepath)
    # return
    return send_file(
        return_file, as_attachment=True, download_name="Template_Rendered.docx"
    )


@app.route("/instruction")
def instructionView():
    return render_template("instruction.html")
